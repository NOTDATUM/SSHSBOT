from openpyxl import load_workbook, Workbook
import random


c_call = 1
c_return = 2

wb = load_workbook("chatDB.xlsx")
ws = wb.active


def loadFile():
    wb = load_workbook("chatDB.xlsx")
    ws = wb.active


def saveFile():
    wb.save("chatDB.xlsx")
    wb.close()
