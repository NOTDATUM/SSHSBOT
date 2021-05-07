from openpyxl import load_workbook, Workbook


wb = load_workbook("shopDB.xlsx")
ws = wb.active


def loadFile():
    wb = load_workbook("shopDB.xlsx")
    ws = wb.active


def saveFile():
    wb.save("shopDB.xlsx")
    wb.close()


def finditem(_name):
    loadFile()
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value == '_name':
            return row
    saveFile()
    return None
